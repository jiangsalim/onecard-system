from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import date, time
from core.models import Student
from attendance.models import Attendance, MealLog
from movement.models import MovementLog
from fees.models import FeeStructure
from core.services import get_payment_balance
from core.mobile_utils import render_mobile_or_desktop
from notifications.models import NotificationSetting
import openpyxl
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def build_pdf_response(filename, elements):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_excel_response(filename, sheet_name, headers, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in data:
        ws.append(row)
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def attendance_report(request):
    """Attendance report with chart — includes Late tracking."""
    today = date.today()
    
    class_filter = request.GET.get('class', '')
    stream_filter = request.GET.get('stream', '')
    
    # Get late cutoff time from settings
    notif_settings = NotificationSetting.objects.first()
    cutoff_time = notif_settings.late_cutoff_time if notif_settings else time(8, 0)
    
    if class_filter:
        from core.services import fetch_students_from_existing_db
        all_school = fetch_students_from_existing_db()
        matching_admissions = [
            s['admission_number'] for s in all_school 
            if s['current_class'] == class_filter 
            and (not stream_filter or s['stream'] == stream_filter)
        ]
        filtered_student_ids = Student.objects.filter(
            admission_number__in=matching_admissions, status='active'
        ).values_list('id', flat=True)
        
        total = len(matching_admissions)
        attendance_list = Attendance.objects.filter(
            scan_date=today, student_id__in=filtered_student_ids
        ).select_related('student').order_by('time_in')
        present = Attendance.objects.filter(scan_date=today, student_id__in=filtered_student_ids).count()
        late = Attendance.objects.filter(scan_date=today, student_id__in=filtered_student_ids, time_in__gt=cutoff_time).count()
    else:
        total = Student.objects.filter(status='active').count()
        attendance_list = Attendance.objects.filter(scan_date=today).select_related('student').order_by('time_in')
        present = Attendance.objects.filter(scan_date=today).count()
        late = Attendance.objects.filter(scan_date=today, time_in__gt=cutoff_time).count()
    
    on_time = present - late
    absent = total - present
    stats = {
        'total': total, 'present': on_time, 'absent': absent, 'late': late,
        'rate': round((present / total * 100) if total > 0 else 0, 1),
    }
    return render_mobile_or_desktop(request, 'reports/attendance.html', 'mobile/reports_attendance.html', {
        'attendance_list': attendance_list, 'stats': stats,
        'class_filter': class_filter, 'stream_filter': stream_filter, 'cutoff_time': cutoff_time,
    })


@login_required
def export_attendance(request):
    fmt = request.GET.get('format', 'xlsx')
    today = date.today()
    records = Attendance.objects.filter(scan_date=today).select_related('student')
    if fmt == 'pdf':
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"Attendance Report — {today}", styles['Title']), Spacer(1, 20)]
        table_data = [['Student ID', 'Admission #', 'Time In', 'Location', 'Marked By']]
        for a in records:
            table_data.append([a.student.id, a.student.admission_number, str(a.time_in), a.scan_location, a.marked_by])
        t = Table(table_data)
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                               ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        elements.append(t)
        return build_pdf_response(f'attendance_report_{today}.pdf', elements)
    else:
        headers = ['Student ID', 'Admission #', 'Date', 'Time In', 'Location', 'Marked By']
        data = [[a.student.id, a.student.admission_number, str(a.scan_date), str(a.time_in), a.scan_location, a.marked_by] for a in Attendance.objects.select_related('student').all()]
        return build_excel_response(f'attendance_report_{today}.xlsx', 'Attendance', headers, data)


@login_required
def export_fees(request):
    fmt = request.GET.get('format', 'xlsx')
    today = date.today()
    students = Student.objects.filter(status='active')
    rows = []
    for s in students:
        paid = get_payment_balance(s.payment_code)
        fee = FeeStructure.objects.first()
        total = float(fee.total_fees) if fee else 800000
        balance = total - float(paid)
        status = 'CLEARED' if balance <= 0 else 'NOT CLEARED'
        rows.append([s.id, s.admission_number, s.payment_code, total, float(paid), balance, status])
    if fmt == 'pdf':
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"Fee Report — {today}", styles['Title']), Spacer(1, 20)]
        table_data = [['Student ID', 'Admission #', 'Payment Code', 'Total', 'Paid', 'Balance', 'Status']] + rows
        t = Table(table_data)
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                               ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        elements.append(t)
        return build_pdf_response(f'fee_report_{today}.pdf', elements)
    else:
        return build_excel_response(f'fee_report_{today}.xlsx', 'Fees', ['Student ID', 'Admission #', 'Payment Code', 'Total', 'Paid', 'Balance', 'Status'], rows)


@login_required
def export_movement(request):
    fmt = request.GET.get('format', 'xlsx')
    today = date.today()
    logs = MovementLog.objects.select_related('student').all()
    rows = []
    for m in logs:
        rows.append([m.student.id, str(m.exit_date), str(m.time_out), str(m.time_in or 'Still Out'), m.get_reason_display(), m.authorized_by])
    if fmt == 'pdf':
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"Movement Log — {today}", styles['Title']), Spacer(1, 20)]
        table_data = [['Student ID', 'Date', 'Time Out', 'Time In', 'Reason', 'Authorized']] + rows
        t = Table(table_data)
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
                               ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                               ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)]))
        elements.append(t)
        return build_pdf_response(f'movement_log_{today}.pdf', elements)
    else:
        return build_excel_response(f'movement_log_{today}.xlsx', 'Movement', ['Student ID', 'Date', 'Time Out', 'Time In', 'Reason', 'Authorized'], rows)


@login_required
def meal_report(request):
    """Meal tracking report — filtered by class if teacher."""
    today = date.today()
    meal_type = request.GET.get('meal', 'lunch')
    class_filter = request.GET.get('class', '')
    stream_filter = request.GET.get('stream', '')
    
    if request.user.role == 'class_teacher':
        class_filter = request.user.assigned_class
        stream_filter = request.user.assigned_stream
    
    meals = MealLog.objects.filter(meal_date=today, meal_type=meal_type).select_related('student')
    
    if class_filter:
        from core.services import fetch_students_from_existing_db
        all_school = fetch_students_from_existing_db()
        matching_admissions = [
            s['admission_number'] for s in all_school 
            if s['current_class'] == class_filter 
            and (not stream_filter or s['stream'] == stream_filter)
        ]
        class_student_ids = Student.objects.filter(admission_number__in=matching_admissions).values_list('id', flat=True)
        meals = meals.filter(student_id__in=class_student_ids)
        total = len(matching_admissions)
    else:
        total = Student.objects.filter(status='active').count()
    
    breakfast_count = MealLog.objects.filter(meal_date=today, meal_type='breakfast').count()
    lunch_count = MealLog.objects.filter(meal_date=today, meal_type='lunch').count()
    supper_count = MealLog.objects.filter(meal_date=today, meal_type='supper').count()

    return render_mobile_or_desktop(request, 'reports/meals.html', 'mobile/reports_meals.html', {
        'meals': meals, 'meal_type': meal_type, 'today': today, 'total': total,
        'served': meals.count(), 'breakfast_count': breakfast_count,
        'lunch_count': lunch_count, 'supper_count': supper_count,
    })